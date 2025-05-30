import os
import re
import unicodedata
import pandas as pd
from flask import Flask, request, render_template, send_file
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx'}

def eh_arquivo_valido(arquivo):
    filename = secure_filename(arquivo.filename)
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def remover_acentos(texto):
    texto = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in texto if not unicodedata.combining(c)).replace('ç', 'c').replace('Ç', 'C')

def separar_num_letra(texto):
    texto = re.sub(r'(?<=\d)(?=[a-zA-Z])', ' ', texto)
    texto = re.sub(r'(?<=[a-zA-Z])(?=\d)', ' ', texto)
    return texto

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        arquivo = request.files.get('file')
        caracteres = request.form.get('caracteres', '.,;:!?@#$%^&*_+=|\\/<>[]{}()-\'"`~')
        minusculo = request.form.get('minusculo') == 'on'
        remover_especiais = request.form.get('remover_especiais') == 'on'
        remover_espacos = request.form.get('remover_espacos') == 'on'

        if not arquivo or not eh_arquivo_valido(arquivo):
            return "Erro: apenas arquivos .xlsx válidos são permitidos.", 400

        try:
            arquivo_bytes = arquivo.read()
            wb = load_workbook(filename=BytesIO(arquivo_bytes), read_only=True)
            ws = wb.active

            data = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header = list(row)
                else:
                    data.append(list(row))
                if i >= 1000:
                    break

            df = pd.DataFrame(data, columns=header)

        except Exception as e:
            return f"Erro ao ler o arquivo Excel: {str(e)}", 400

        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str)

                if minusculo:
                    df[col] = df[col].str.lower()

                if remover_especiais:
                    padrao_regex = f"[{''.join(re.escape(c) for c in caracteres)}]"
                    df[col] = df[col].apply(lambda x: re.sub(padrao_regex, ' ', x))
                    df[col] = df[col].apply(remover_acentos)

                df[col] = df[col].apply(separar_num_letra)

                if remover_espacos:
                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True).str.strip()

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='dados_processados.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('index.html', max_size_mb=10)

@app.errorhandler(413)
def too_large(e):
    return "Erro: o arquivo excede o tamanho máximo permitido (10 MB).", 413

if __name__ == '__main__':
    app.run(debug=True)
